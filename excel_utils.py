import os
import shutil
from typing import List, Dict, Any, Final
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from datetime import datetime, date
from pathlib import Path

# Constantes de directorios y límites
ARCHIVOS_DIR = "archivos"
PLANTILLA_PATH = "Plantillas/plantilla_base"
MAX_FILA_PERMITIDA = 300 # Fila máxima permitida para datos (hasta 219)
FILA_ENCABEZADO = 10 

# Constantes de columnas
COL_NRO = 1          # A: Número de fila
COL_RADICADO = 2     # B: Radicado
COL_TIPO = 3         # C: Tipo de audiencia
COL_FECHA = 4        # D: Fecha
COL_HORA = 5         # E: Hora
COL_JUZGADO = 6      # F: Juzgado
COL_REALIZADO_SI = 7      # G
COL_REALIZADO_NO = 8      # H
COL_MOTIVOS_INICIO = 9    # I
COL_MOTIVOS_FIN = 16      # P
COL_OBSERVACIONES = 17    # Q

# Nueva constante global con los tipos de audiencia válidos
TIPOS_AUDIENCIA_VALIDOS: Final[List[str]] = [
    "Alegatos de conclusión",
    "Audiencia concentrada",
    "Audiencia de acusación",
    "Audiencia de conciliación",
    "Audiencia de control de legalidad",
    "Audiencia de individualización de pena",
    "Audiencia de imputación",
    "Audiencia de incidente de reparación integral",
    "Audiencia de juicio oral",
    "Audiencia de medidas de aseguramiento",
    "Audiencia de nulidad",
    "Audiencia de preclusión",
    "Audiencia de prórroga",
    "Audiencia de revisión de medida",
    "Audiencia de verificación de cumplimiento",
    "Audiencia preliminar",
    "Audiencia preparatoria",
    "Otra"
]

def crear_copia_plantilla(nombre_archivo: str) -> str:
    """
    Copia la plantilla base a un nuevo archivo de trabajo.
    
    Args:
        nombre_archivo: Nombre para el nuevo archivo.
    
    Returns:
        str: Ruta del nuevo archivo creado.
    
    Raises:
        FileNotFoundError: Si no se encuentra la plantilla base.
        ValueError: Si el archivo destino ya existe.
    """
    if not os.path.exists(PLANTILLA_PATH):
        raise FileNotFoundError(
            f"No se encontró la plantilla base en '{PLANTILLA_PATH}'"
        )

    if not nombre_archivo.endswith('.xlsx'):
        nombre_archivo += '.xlsx'
        
    destino = os.path.join(ARCHIVOS_DIR, nombre_archivo)
    if os.path.exists(destino):
        raise ValueError(
            f"Ya existe un archivo con el nombre '{nombre_archivo}'. "
            "Elija otro nombre."
        )
    
    os.makedirs(ARCHIVOS_DIR, exist_ok=True)
    shutil.copy2(PLANTILLA_PATH, destino)
    return destino

def listar_archivos():
    """
    Lista los archivos .xlsx en la carpeta archivos/.
    """
    if not os.path.exists(ARCHIVOS_DIR):
        return []
    return [f for f in os.listdir(ARCHIVOS_DIR) if f.endswith('.xlsx')]

def limpiar_celdas_combinadas(ws: Worksheet) -> None:
    """
    Descombina todas las celdas combinadas en la hoja.
    """
    if ws.merged_cells.ranges:
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))

def parse_fecha_hora(d: Dict[str, Any]) -> datetime:
    """
    Parsea y valida los campos 'fecha' y 'hora' en formato dd/mm/yyyy y HH:MM.
    Lanza ValueError si el formato es inválido.
    """
    fecha = d.get('fecha')
    hora = d.get('hora')
    if not fecha or not hora:
        raise ValueError("Los campos 'fecha' y 'hora' son obligatorios.")
    try:
        dt = datetime.strptime(fecha + " " + hora, "%d/%m/%Y %H:%M")
    except Exception:
        raise ValueError(f"Fecha y hora inválidas: '{fecha} {hora}'. Formato esperado: dd/mm/yyyy HH:MM")
    return dt

def validar_campos_audiencia(d: Dict[str, Any]) -> None:
    """
    Verifica que los campos obligatorios estén presentes y válidos.
    
    Args:
        d: Diccionario con los datos de la audiencia
    
    Raises:
        ValueError: Si falta algún campo obligatorio o si los valores no son válidos
    """
    obligatorios = ['radicado', 'tipo_audiencia', 'fecha', 'hora', 'juzgado', 'se_realizo']
    for campo in obligatorios:
        if not d.get(campo):
            raise ValueError(f"El campo '{campo}' es obligatorio.")
    
    # Valida tipo_audiencia contra la lista de tipos válidos
    tipo_audiencia = str(d['tipo_audiencia']).strip()
    if tipo_audiencia not in TIPOS_AUDIENCIA_VALIDOS:
        raise ValueError(
            f"El tipo de audiencia '{tipo_audiencia}' no es válido. "
            f"Debe ser uno de los siguientes valores: {', '.join(TIPOS_AUDIENCIA_VALIDOS)}"
        )
    
    # Valida se_realizo
    se_realizo = str(d['se_realizo']).strip().upper()
    if se_realizo not in ['SI', 'NO']:
        raise ValueError("El campo 'se_realizo' debe ser 'SI' o 'NO'.")
    d['se_realizo'] = se_realizo  # Normaliza el valor

    # Normaliza el tipo de audiencia para asegurar consistencia
    d['tipo_audiencia'] = tipo_audiencia

def guardar_audiencias_excel(
    datos: List[Dict[str, Any]],
    nombre_archivo: str
) -> None:
    """
    Guarda una lista de audiencias en el archivo Excel especificado.
    Aplica estilos a las celdas de totales y motivos.
    """
    # Valida que no se esté usando la plantilla directamente
    validar_no_es_plantilla(nombre_archivo)
    
    ruta = os.path.join(ARCHIVOS_DIR, nombre_archivo)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"El archivo {ruta} no existe.")

    wb = load_workbook(ruta)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("No se pudo cargar la hoja activa del archivo Excel.")
    limpiar_celdas_combinadas(ws)

    # Borra todas las filas de datos previas, pero preserva la firma
    max_row = min(ws.max_row, MAX_FILA_PERMITIDA - 1)  # Nunca borrar más allá de fila 188
    if max_row > FILA_ENCABEZADO:
        filas_a_borrar = max_row - FILA_ENCABEZADO
        ws.delete_rows(FILA_ENCABEZADO + 1, filas_a_borrar)

    # Validación y ordenamiento
    for d in datos:
        validar_campos_audiencia(d)
        parse_fecha_hora(d)

    datos_ordenados = sorted(datos, key=parse_fecha_hora, reverse=True)

    # Define estilos para las celdas
    borde_fino = Side(style='thin')
    bordes_celda = Border(
        left=borde_fino,
        right=borde_fino,
        top=borde_fino,
        bottom=borde_fino
    )
    
    fondo_totales = PatternFill(
        start_color='E0E0E0',  # Gris claro
        end_color='E0E0E0',
        fill_type='solid'
    )

    # Escribe los datos y aplica estilos
    for idx, d in enumerate(datos_ordenados, start=1):
        fila = FILA_ENCABEZADO + idx
        
        # Escribe datos
        ws.cell(row=fila, column=COL_NRO, value=idx)
        ws.cell(row=fila, column=COL_RADICADO, value=d["radicado"])
        ws.cell(row=fila, column=COL_TIPO, value=d["tipo_audiencia"])
        ws.cell(row=fila, column=COL_FECHA, value=d["fecha"])
        ws.cell(row=fila, column=COL_HORA, value=d["hora"])
        ws.cell(row=fila, column=COL_JUZGADO, value=d["juzgado"])
        
        # Escribe SI/NO como X en la columna correspondiente
        ws.cell(row=fila, column=COL_REALIZADO_SI, value="")  # Limpia
        ws.cell(row=fila, column=COL_REALIZADO_NO, value="")  # Limpia
        if d["se_realizo"] == "SI":
            ws.cell(row=fila, column=COL_REALIZADO_SI, value="X")
        elif d["se_realizo"] == "NO":
            ws.cell(row=fila, column=COL_REALIZADO_NO, value="X")

        ws.cell(row=fila, column=COL_OBSERVACIONES, value=d.get("observaciones", ""))
        # Motivos
        for i, col in enumerate(range(COL_MOTIVOS_INICIO, COL_MOTIVOS_FIN + 1)):
            motivo = d.get("motivos", [])[i] if i < len(d.get("motivos", [])) else ""
            ws.cell(row=fila, column=col, value=motivo)
            ws.cell(row=fila, column=col, value=motivo)

        # Aplica bordes a las celdas de motivos
        for col in range(COL_MOTIVOS_INICIO, COL_MOTIVOS_FIN + 1):
            celda = ws.cell(row=fila, column=col)
            celda.border = bordes_celda

        # Copia estilos de la fila 11
        copiar_estilos_fila(ws, FILA_ENCABEZADO + 1, fila)
    
    # --- ELIMINA FILAS DE TOTALES PREVIAS ---
    # Busca desde la fila justo después de los datos hacia abajo
    fila_datos_fin = FILA_ENCABEZADO + len(datos_ordenados)
    fila_busqueda = fila_datos_fin + 1
    filas_a_borrar = []
    while ws.cell(row=fila_busqueda, column=7).value or ws.cell(row=fila_busqueda, column=8).value:
        valor_g = str(ws.cell(row=fila_busqueda, column=7).value or "")
        valor_h = str(ws.cell(row=fila_busqueda, column=8).value or "")
        if valor_g.startswith("TOTAL DE") or valor_h.startswith("TOTAL DE"):
            filas_a_borrar.append(fila_busqueda)
            fila_busqueda += 1
        else:
            break
    # Borra de abajo hacia arriba para no desplazar filas
    for f in reversed(filas_a_borrar):
        ws.delete_rows(f, 1)

    # Agrega nueva fila de totales
    total_si = sum(1 for d in datos_ordenados if d["se_realizo"] == "SI")
    total_no = sum(1 for d in datos_ordenados if d["se_realizo"] == "NO")
    fila_totales = FILA_ENCABEZADO + len(datos_ordenados) + 1

    if fila_totales > MAX_FILA_PERMITIDA:
        raise ValueError("Demasiadas audiencias: podrías sobrescribir la firma del defensor.")

    # Escribe totales en columnas G y H
    ws.cell(row=fila_totales, column=COL_REALIZADO_SI, 
            value=f"TOTAL DE AUDIENCIAS REALIZADAS: {total_si}")
    ws.cell(row=fila_totales, column=COL_REALIZADO_NO, 
            value=f"TOTAL DE AUDIENCIAS NO REALIZADAS: {total_no}")

    # Aplica estilos a las celdas de totales
    celda_total_si = ws.cell(row=fila_totales, column=COL_REALIZADO_SI)
    celda_total_no = ws.cell(row=fila_totales, column=COL_REALIZADO_NO)

    celda_total_si.fill = fondo_totales
    celda_total_no.fill = fondo_totales
    celda_total_si.border = bordes_celda
    celda_total_no.border = bordes_celda

    # Aplica estilos a totales
    aplicar_estilos_totales(ws.cell(row=fila_totales, column=COL_REALIZADO_SI))
    aplicar_estilos_totales(ws.cell(row=fila_totales, column=COL_REALIZADO_NO))

    # Calcula la fila para totales de motivos
    fila_totales_motivos = fila_totales + 1
    
    # Verifica que no sobrepase el límite permitido
    if fila_totales_motivos >= MAX_FILA_PERMITIDA:
        raise ValueError(
            "No hay espacio suficiente para escribir los totales de motivos "
            "sin sobrescribir la firma del defensor."
        )

    # Mapeo de columnas a nombres de motivos
    motivos_nombres = {
        COL_MOTIVOS_INICIO: "Juez",
        COL_MOTIVOS_INICIO + 1: "Fiscalía",
        COL_MOTIVOS_INICIO + 2: "Usuario",
        COL_MOTIVOS_INICIO + 3: "Inpec",
        COL_MOTIVOS_INICIO + 4: "Víctima",
        COL_MOTIVOS_INICIO + 5: "ICBF",
        COL_MOTIVOS_INICIO + 6: "Defensor Confianza",
        COL_MOTIVOS_INICIO + 7: "Defensor Público"
    }

    # Cuenta ocurrencias de cada motivo
    conteo_motivos = {col: 0 for col in range(COL_MOTIVOS_INICIO, COL_MOTIVOS_FIN + 1)}
    
    # Solo cuenta motivos de audiencias no realizadas
    for fila in range(FILA_ENCABEZADO + 1, fila_totales):
        # Verifica si la audiencia no se realizó
        if ws.cell(row=fila, column=COL_REALIZADO_NO).value == "X":
            for col in conteo_motivos.keys():
                if ws.cell(row=fila, column=col).value:
                    conteo_motivos[col] += 1

    # Escribe los totales de motivos y aplica estilos
    for col, total in conteo_motivos.items():
        # Escribe el total
        celda_total = ws.cell(
            row=fila_totales_motivos, 
            column=col,
            value=f"{motivos_nombres[col]}: {total}"
        )
        # Aplica estilos
        aplicar_estilos_totales_motivos(celda_total)

    # Guarda y cierra el archivo
    wb.save(ruta)
    wb.close()

def guardar_una_audiencia_excel(
    d: Dict[str, Any], 
    nombre_archivo: str
) -> Dict[str, Any]:
    """
    Guarda una sola audiencia en el archivo Excel.
    Retorna un diccionario con el estado y número de audiencias guardadas.
    """
    ruta = os.path.join(ARCHIVOS_DIR, nombre_archivo)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"El archivo {ruta} no existe.")

    wb = load_workbook(ruta)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("No se pudo cargar la hoja activa del archivo Excel.")

    # Leer audiencias existentes
    audiencias_existentes = []
    fila = FILA_ENCABEZADO + 1
    while ws.cell(row=fila, column=COL_RADICADO).value:
        # Determina se_realizo basado en X en columnas G/H
        valor_si = ws.cell(row=fila, column=COL_REALIZADO_SI).value
        valor_no = ws.cell(row=fila, column=COL_REALIZADO_NO).value
        se_realizo = ""
        if valor_si and str(valor_si).strip() == "X":
            se_realizo = "SI"
        elif valor_no and str(valor_no).strip() == "X":
            se_realizo = "NO"

        registro = {
            "radicado": ws.cell(row=fila, column=COL_RADICADO).value,
            "tipo_audiencia": ws.cell(row=fila, column=COL_TIPO).value,
            "fecha": ws.cell(row=fila, column=COL_FECHA).value,
            "hora": ws.cell(row=fila, column=COL_HORA).value,
            "juzgado": ws.cell(row=fila, column=COL_JUZGADO).value,
            "se_realizo": se_realizo,
            "observaciones": ws.cell(row=fila, column=COL_OBSERVACIONES).value,
            "motivos": [
                ws.cell(row=fila, column=col).value or ""
                for col in range(COL_MOTIVOS_INICIO, COL_MOTIVOS_FIN + 1)
            ]
        }
        audiencias_existentes.append(registro)
        fila += 1

    # Verifica duplicados
    nuevo_radicado = str(d.get('radicado', '')).strip()
    if any(str(a.get('radicado', '')).strip() == nuevo_radicado 
           for a in audiencias_existentes):
        wb.close()
        raise ValueError(f"Ya existe una audiencia con el radicado '{nuevo_radicado}'.")

    # Agrega y guarda
    audiencias_existentes.append(d)
    wb.close()
    guardar_audiencias_excel(audiencias_existentes, nombre_archivo)

    return {
        "estado": "ok",
        "audiencias_guardadas": len(audiencias_existentes)
    }

def validar_no_es_plantilla(nombre_archivo: str) -> None:
    """
    Verifica que el archivo a usar no sea la plantilla base.
    Lanza ValueError si se intenta usar la plantilla directamente.
    """
    ruta_absoluta = os.path.abspath(os.path.join(ARCHIVOS_DIR, nombre_archivo))
    plantilla_absoluta = os.path.abspath(PLANTILLA_PATH)
    
    if ruta_absoluta == plantilla_absoluta:
        raise ValueError(
            "No se puede modificar la plantilla base directamente. "
            "Use crear_copia_plantilla() para crear una copia de trabajo."
        )

def exportar_con_firma(nombre_archivo: str) -> str:
    """
    Crea una copia del archivo con firma del defensor.
    Aplica estilos a la fila de firma.
    """
    # Validaciones iniciales
    if not nombre_archivo.endswith('.xlsx'):
        nombre_archivo += '.xlsx'
    
    ruta_origen = os.path.join(ARCHIVOS_DIR, nombre_archivo)
    if not os.path.exists(ruta_origen):
        raise FileNotFoundError(f"No se encontró el archivo {ruta_origen}")
    
    # Genera nombre del archivo exportado
    nombre_base = Path(nombre_archivo).stem
    fecha_actual = date.today().strftime("%Y%m%d")
    nombre_exportado = f"{nombre_base}_exportado_{fecha_actual}.xlsx"
    ruta_destino = os.path.join(ARCHIVOS_DIR, nombre_exportado)
    
    # Copia el archivo
    shutil.copy2(ruta_origen, ruta_destino)
    
    # Abre la copia y agrega la firma
    wb = load_workbook(ruta_destino)
    ws = wb.active
    if ws is None:
        wb.close()
        os.unlink(ruta_destino)  # Elimina la copia si hay error
        raise ValueError("No se pudo procesar el archivo Excel")

    # Encuentra la última fila con datos incluyendo totales y motivos
    ultima_fila = FILA_ENCABEZADO
    while ws.cell(row=ultima_fila + 1, column=COL_RADICADO).value:
        ultima_fila += 1
    
    # Avanza más allá de los totales y motivos
    while (ws.cell(row=ultima_fila + 1, column=COL_REALIZADO_SI).value or 
           ws.cell(row=ultima_fila + 1, column=COL_REALIZADO_NO).value):
        ultima_fila += 1
    
    # Inserta la firma en la siguiente fila
    fila_firma = ultima_fila + 1
    
    # Define estilos para la firma
    borde_grueso = Side(style='thick')
    borde_firma = Border(
        left=borde_grueso,
        right=borde_grueso,
        top=borde_grueso,
        bottom=borde_grueso
    )
    
    alineacion_centro = Alignment(
        horizontal='center',
        vertical='center'
    )

    # Aplica estilos a la celda combinada de firma
    celda_firma = ws[f"A{fila_firma}"]
    celda_firma.value = "Firma del defensor público:__________________________"
    celda_firma.alignment = alineacion_centro
    
    # Combina celdas y aplica borde exterior
    rango_firma = f"A{fila_firma}:Q{fila_firma}"
    ws.merge_cells(rango_firma)
    
    # Aplica el borde al rango completo
    for col in range(1, 18):  # A hasta Q
        celda = ws.cell(row=fila_firma, column=col)
        if col == 1:  # Primera celda
            celda.border = borde_firma
        else:
            celda.border = Border(top=borde_grueso, bottom=borde_grueso)
            if col == 17:  # Última celda
                celda.border = Border(
                    top=borde_grueso,
                    bottom=borde_grueso,
                    right=borde_grueso
                )

    # Guarda y cierra
    wb.save(ruta_destino)
    wb.close()
    
    return ruta_destino

def copiar_estilos_fila(
    ws: Worksheet,
    fila_origen: int,
    fila_destino: int
) -> None:
    """
    Copia los estilos de una fila origen a una fila destino.
    
    Args:
        ws: Hoja de trabajo activa
        fila_origen: Número de fila desde donde copiar estilos
        fila_destino: Número de fila donde aplicar los estilos
    """
    # Define estilos base por si no se pueden obtener de la fila origen
    fuente_base = Font(
        name='Calibri',
        size=11,
        color='000000',
        bold=False,
        italic=False
    )
    
    alineacion_base = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    relleno_base = PatternFill(
        fill_type='solid',
        fgColor='FFFFFF'
    )
    
    borde_fino = Side(
        style='thin',
        color='000000'
    )
    
    bordes_base = Border(
        top=borde_fino,
        bottom=borde_fino,
        left=borde_fino,
        right=borde_fino
    )
    
    # Obtiene altura de la fila origen o usa valor por defecto
    altura_fila: float = ws.row_dimensions[fila_origen].height or 20.0
    ws.row_dimensions[fila_destino].height = altura_fila
    # Copia estilos celda por celda
    for col in range(1, 18):  # A hasta Q
        celda_origen = ws.cell(row=fila_origen, column=col)
        celda_destino = ws.cell(row=fila_destino, column=col)
        
        # Crea una nueva celda con los estilos combinados
        ws._cells[(fila_destino, col)] = Cell(
            ws,
            row=fila_destino,
            column=col,
            value=celda_destino.value
        )
        celda_destino = ws.cell(row=fila_destino, column=col)
        
        # Copia estilos individuales creando nuevos objetos
        origen_font = celda_origen.font
        celda_destino.font = Font(
            name=origen_font.name,
            size=origen_font.size,
            bold=origen_font.bold,
            italic=origen_font.italic,
            color=origen_font.color
        )
        
        origen_fill = celda_origen.fill
        celda_destino.fill = PatternFill(
            fill_type=origen_fill.fill_type,
            start_color=origen_fill.start_color,
            end_color=origen_fill.end_color
        )
        
        origen_alignment = celda_origen.alignment
        celda_destino.alignment = Alignment(
            horizontal=origen_alignment.horizontal,
            vertical=origen_alignment.vertical,
            wrap_text=origen_alignment.wrap_text
        )
        
        celda_destino.number_format = celda_origen.number_format
        # Create new Border object from original or default
        origen_border = celda_origen.border or bordes_base
        celda_destino.border = Border(
            left=origen_border.left,
            right=origen_border.right,
            top=origen_border.top,
            bottom=origen_border.bottom
        )
        
        # Establece ancho de columna si no está definido
        letra_col = chr(64 + col)  # Convierte número a letra (1=A, 2=B, etc.)
        if not ws.column_dimensions[letra_col].width:
            ancho_origen = ws.column_dimensions[letra_col].width
            ws.column_dimensions[letra_col].width = ancho_origen or 15.0

from typing import Union
from openpyxl.cell.cell import MergedCell

def aplicar_estilos_totales(celda: Union["Cell", "MergedCell"]) -> None:
    """Aplica estilos específicos a las celdas de totales."""
    borde_fino = Side(style='thin', color='000000')
    celda.fill = PatternFill(
        fill_type='solid',
        fgColor='D9D9D9'
    )
    celda.border = Border(
        top=borde_fino,
        bottom=borde_fino,
        left=borde_fino,
        right=borde_fino
    )
    celda.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

def aplicar_estilos_totales_motivos(celda: Union[Cell, MergedCell]) -> None:
    """
    Aplica estilos específicos a las celdas de totales de motivos.
    Sin color de fondo, solo bordes finos y alineación centrada.
    """
    borde_fino = Side(style='thin', color='000000')
    celda.border = Border(
        top=borde_fino,
        bottom=borde_fino,
        left=borde_fino,
        right=borde_fino
    )
    celda.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    celda.font = Font(
        name='Calibri',
        size=11,
        color='000000'
    )

def aplicar_estilos_firma(ws: Worksheet, fila: int) -> None:
    """Aplica estilos específicos a la fila de firma."""
    borde_grueso = Side(style='thick', color='000000')
    
    # Aplica bordes gruesos al rango completo
    for col in range(1, 18):  # A hasta Q
        celda = ws.cell(row=fila, column=col)
        if col == 1:  # Primera celda
            celda.border = Border(
                left=borde_grueso,
                right=Side(style=None),
                top=borde_grueso,
                bottom=borde_grueso
            )
        elif col == 17:  # Última celda
            celda.border = Border(
                left=Side(style=None),
                right=borde_grueso,
                top=borde_grueso,
                bottom=borde_grueso
            )
        else:  # Celdas intermedias
            celda.border = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=borde_grueso,
                bottom=borde_grueso
            )

# Ejemplo actualizado de uso en FastAPI:
# -----------------------------------
# from fastapi import FastAPI, HTTPException
# from pydantic import BaseModel
# from typing import List
# import excel_utils
#
# app = FastAPI()
#
# class Audiencia(BaseModel):
#     radicado: str
#     tipo_audiencia: str
#     fecha: str
#     hora: str
#     juzgado: str
#     se_realizo: str
#     motivos: List[str] = []
#     observaciones: str = ""
#     nombre_archivo: str
#
# @app.post("/audiencias/")
# def guardar_audiencia(audiencia: Audiencia):
#     try:
#         excel_utils.guardar_una_audiencia_excel(
#             audiencia.dict(),
#             audiencia.nombre_archivo
#         )
#         return {"ok": True}
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=str(e))
#
# @app.post("/crear_archivo/")
# def crear_archivo(nombre: str):
#     try:
#         ruta = excel_utils.crear_copia_plantilla(nombre)
#         return {"ruta": ruta}
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=str(e))
#
# @app.get("/archivos/")
# def listar():
#     return excel_utils.listar_archivos()
# --- FastAPI endpoint example (moved here for clarity) ---
from fastapi import FastAPI, HTTPException

app = FastAPI()

import excel_utils as excel_utils_module

@app.post("/exportar/{nombre_archivo}")
def exportar_archivo(nombre_archivo: str):
    try:
        ruta_exportado = excel_utils_module.exportar_con_firma(nombre_archivo)
        return {"archivo_exportado": ruta_exportado}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))